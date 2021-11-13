from datetime import datetime
from openpyxl import (
    Workbook,
    drawing
)
from openpyxl.worksheet.table import Table, TableStyleInfo
import os


class ExcelManager:
    def __init__(self, datas, export_name):
        self.datas = datas
        self.export_name = export_name

    def export(self):
        # create a new workbook
        wb = Workbook()
        wb.remove(wb["Sheet"])

        # ------------------worksheet: overall------------------
        wb.create_sheet("overall")
        ws = wb["overall"]
        ws.append(("Date: ", datetime.now().strftime('%Y/%m/%d %H:%M:%S')))

        # blank line
        ws.append([])

        # compressive stress
        ws.append(
            ("Average Compressive Stress [f'c(kgf/cm^2)]",))
        ws.append(
            ["Name", *[i+1 for i in range(max(self.datas["exNum"]))], "Average"])
        max_stress_list = []
        for i in range(len(self.datas["exNum_ids"])-1):
            each_row = []
            each_row += [self.datas["ratio_type"][i]]
            each_row += self.datas["max_stress_list"][self.datas["exNum_ids"][i]:self.datas["exNum_ids"][i+1]]
            max_stress_list.append(each_row)
        for row in max_stress_list:
            ws.append(row)

        column = max(self.datas["exNum"]) + 2
        row = 5
        for i, value in enumerate(self.datas["avg_max_stress_list"]):
            ws.cell(column=column, row=row+i, value=value)

        # blank line
        ws.append([])

        # elastic modulus
        ws.append(("Average Elastic Modulus [f'c(kgf/cm^2)]",))
        ws.append(["Name", *[i+1 for i in range(max(self.datas["exNum"]))], "Average"])
        max_em_list = []
        for i in range(len(self.datas["exNum_ids"])-1):
            each_row = []
            each_row += [self.datas["ratio_type"][i]]
            each_row += self.datas["max_em_list"][self.datas["exNum_ids"][i]:self.datas["exNum_ids"][i+1]]
            max_em_list.append(each_row)
        for row in max_em_list:
            ws.append(row)

        column = max(self.datas["exNum"]) + 2
        row = 8 + len(self.datas["ratio_type"])
        for i, value in enumerate(self.datas["avg_max_em_list"]):
            ws.cell(column=column, row=row+i, value=value)

        # ---------------------detail datas for each specimen-------------------
        for data in self.datas["specimens_datas"]:

            Id = data["id"]
            name = data["name"]
            sheet_name = f"{name}_{str(Id)}"

            wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            ws.append(("id", Id))
            ws.append(("name", name))
            ws.append(tuple(data["info"].keys()))

            for dis, force, d1, d2, strain, stress in zip(
                data["info"]["displacement"]["data"],
                data["info"]["force"]["data"],
                data["info"]["D1"]["data"],
                data["info"]["D2"]["data"],
                data["info"]["strain"]["data"],
                data["info"]["stress"]["data"],
            ):
                ws.append((dis, force, d1, d2, strain, stress))

            ws["G4"] = data["info"]["elastic_modulus"]["data"]["slope"]

            img = drawing.image.Image(os.path.join(self.datas["save_folder_path"], f"figure_{str(Id)}.png"))
            img.anchor = 'J4'
            ws.add_image(img)

        # save workbook
        wb.save(os.path.join(self.datas["save_folder_path"], self.export_name))


if __name__ == "__main__":
    save_folder_path = "test"
    export_name = "test.xlsx"
    ratio_type = ["a", "b", "c"]
    exNum = [8, 8, 8]
    exNum_ids = [0, 8, 16, 24]
    max_stress_list = [487.8782025451012, 394.90680908538394, 470.8807531398036, 479.1992255407472, 455.39291676128704, 460.25987167628807, 526.3171226947065, 502.8338221981192, 491.38918979245034, 332.42386983764555, 396.49677507874475, 492.2086446868025, 522.3591388120803, 499.9975091405196, 515.9022957776968,
                       491.14270567222536, 498.12378334933464, 470.5519692035423, 434.33583197372644, 520.148322526952, 514.355562821273, 493.76328185202084, 450.6243815758245]
    max_em_list = [255102.42946974566, 291237.26143419626, 165363.22138128575, 219797.7786709594, 179219.94736819898, 144130.33501921486, 147198.00145365525, 277882.92673007806, 166214.73640726766, 203378.81615245258, 384800.81900393177,
                   281997.75705631764, 264823.6141058004, 240252.7074908778, 220567.41075029605, 288649.30810198357, 239174.49672636075, 275456.5903138937, 254992.0701377906, 260402.7511353376, 252022.10843067188, 215407.8708265754, 266242.5216168384]
    specimens_datas = []
    ExcelManager(
        save_folder_path,
        export_name,
        ratio_type,
        exNum,
        exNum_ids,
        max_stress_list,
        max_em_list,
        specimens_datas
    ).export()
