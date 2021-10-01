#!/usr/bin/env python
# coding: utf-8

# Author madihsiang

# # Elastic Modulus Test UI 楊氏係數試驗

import pandas as pd
import numpy as np
import seaborn as sns; sns.set()
import matplotlib as mpl
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei'] 
from scipy.optimize import curve_fit
import os
import time
import copy

import re
from PyQt5 import QtWidgets, QtGui
import sys
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from PyQt5.QtCore import QBasicTimer,QAbstractTableModel,Qt

from PyQt5.QtWidgets import (QGraphicsView,QGraphicsScene,QApplication,QGraphicsPixmapItem,QProgressBar)
from PyQt5.QtGui import *
from matplotlib.font_manager import FontProperties
# from matplotlib.figure import Figure
# from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PIL import Image

import xlsxwriter

# %matplotlib inline

# 用一次式 ax+b 迴歸
def func(x,a,b):
    return a*x+b

def read():
    global fileName
    fileName, filetype = QFileDialog.getOpenFileName(w,
                  "選取檔案",
                  "",
                  "txt file(*.txt)", None)
    if fileName == '':
        return
    w.PathTxt.setText(fileName)

def change_row():
    w.inputWidget.setRowCount(w.spinBox.value())

def save():
    global PathFile
    PathFile = QFileDialog.getExistingDirectory(w,
                  "選取資料夾",
                  "")
    if PathFile == '':
        return
    
    w.savePathTxt.setText(PathFile)
    
def compshow():
    w.compLabel.setText(str(w.compVerticalSlider.value())+"%")
    
def EMshow():
    w.EMLabel.setText(str(w.EMVerticalSlider.value())+"%")
    
# 繪圖函式
def axesPlot(ax1,ax2,x,y,i,Strain_i_List,ex,xmax,ymax,popt):
    ax1.grid(ls='-')
    ax1.set(title = 'Data_All',xlabel = 'strain',ylabel = 'f"c(kg/cm2)',xlim = (0,max(Strain_i_List)*1.25))
    ax1.set_label('sine')
    ax1.scatter(x,y)
#             ax1.set_xlim(0,0.003)
    ax1.legend(['{}'.format(ex[i])],loc='upper left')

    ax2.set(title = 'Data_Until_Max',xlabel = 'strain',ylabel = 'f"c(kg/cm2)')
    ax2.grid(ls = '-')
    ax2.plot(xmax,ymax,'o',xmax,func(xmax,popt[0],popt[1]),'r')
#             ax2.set_xlim(0,0.0007)
    ax2.legend(['{}'.format(ex[i]),'Elastic Modulus: {}'.format(round(popt[0],2))],loc='upper left')
    
# 畫直方圖
def createLabels(data):
    for item in data:
        height = item.get_height()
        plt.text(
            item.get_x()+item.get_width()/2., 
            height*1.00, 
            '%.2f' % float(height),
            ha = "center",
            va = "bottom",
        )
        
        
# 放預設預覽圖
def Initial_layout_Figure(lab,text,scene,graphicsView):
    lab.setText(text)
    lab.setStyleSheet("QLabel{color:rgb(0,0,0,255);background-color:rgb(255,255,255,255);font-size:30px;font-weight:normal;font-family:Arial;}")
    lab.setScaledContents(True)

    scene.addWidget(lab)
    graphicsView.setScene(scene)    
        
        
# 放預覽圖
def layout_Figure(lab,figName,scene,graphicsView,scale = [1,1]):
    img = Image.open(figName)
    lab.setFixedWidth(img.size[0]*scale[0])
    lab.setFixedHeight(img.size[1]*scale[1])
    lab.setPixmap(QPixmap(figName))
    lab.setScaledContents(True)

    scene.addWidget(lab)
    graphicsView.setScene(scene)
        
# 修改TableView的header -> column,row
def header(data):
    model = QStandardItemModel()
    model.setHorizontalHeaderLabels(list(map(str,data.keys())))  # map可以直接將list裡面的型態轉型
    model.setVerticalHeaderLabels(list(map(str,data.index)))

    for row in range(data.shape[0]):
        for column in range(data.shape[1]):
            item = QStandardItem()
            item.setText(str(data.iloc[row][column]))
            model.setItem(row, column, item)

    return model


# 重新篩選並繪comp圖
def compRePlot():
    try:
        newComp = copy.copy(comp)  # 複製一份作修改
        for i in range(len(exRatio)):
            delRatio = w.compVerticalSlider.value()/100
            avg = comp_avg.iloc[-1][i]
            newComp[exRatio[i]] = comp[exRatio[i]][(comp[exRatio[i]] < avg*(1+delRatio))&(comp[exRatio[i]] > avg*(1-delRatio))]

        # 新的pd加上新的平均值
        newCompAvg = pd.concat([newComp,pd.DataFrame(newComp.mean()).rename(columns = {0:"平均"}).T])

        compModel = header(newCompAvg)
        w.compTableView_2.setModel(compModel)

        # 畫直方圖 -> 抗壓強度
        if True not in np.isnan(list(newCompAvg.iloc[-1])):
#         if newCompAvg.count()[-1] != 0:
            compAvgList = list(newCompAvg.iloc[-1])
            plt.figure(figsize=(5,3))
            A = plt.bar(exRatio, compAvgList, align='center', alpha=0.5)
            # plt.xticks(y_pos, objects)
            plt.ylabel('Compressive Strength(kgf/cm2)');
            compTitleName = w.compTitle.text()
            if compTitleName == "":
                compTitleName = 'Compressive Strength'
            plt.title(compTitleName)

            createLabels(A)
            plt.savefig(path+'\\Compressive Strength-剔除歧異值.png',bbox_inches ='tight');

            figName = path+"\\Compressive Strength-剔除歧異值.png"
            layout_Figure(lab4,figName,scene4,w.graphicsView_4,scale=[1,1])

            # ToExcel
            ExcelFileName = path+'\\抗壓強度-歧異值.xlsx'
    #         workbook = xlsxwriter.Workbook(ExcelFileName)     #新建文件
    #         worksheet1 = workbook.add_worksheet('Compressive Strength')             #新建sheet
    #         bold = workbook.add_format({'bold': True})
            newCompAvg = newCompAvg.fillna("剔除掉")
            newCompAvg.to_excel(ExcelFileName)
    #         worksheet1.insert_image(1,10, path+'\\Compressive Strength-剔除歧異值.png',{'x_scale':1,'y_scale':1})
    #         workbook.close()  # 保存並關閉

    #             from openpyxl import load_workbook

    #             book = load_workbook(ExcelFileName)
    #             writer = pd.ExcelWriter(ExcelFileName, engine='openpyxl') 
    #             writer.book = book
    #             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    #             newCompAvg.to_excel(writer, "Compressive Strength")

            QMessageBox.information(w,'Congratulation', "剔除後的抗壓數據成功存成Excel~!")
            plt.close("all")

        else:
            QMessageBox.critical(w,'Error', "該剔除值下，有配比已經無資料!")
    except:
        QMessageBox.critical(w,'Error', "尚未有抗壓數據!")


# 重新篩選並繪EM圖
def EMRePlot():
    try:
        newEM = copy.copy(EM)  # 複製一份作修改
        for i in range(len(exRatio)):
            delRatio = w.EMVerticalSlider.value()/100
            avg = EM_avg.iloc[-1][i]
            newEM[exRatio[i]] = EM[exRatio[i]][(EM[exRatio[i]] < avg*(1+delRatio))&(EM[exRatio[i]] > avg*(1-delRatio))]

        # 新的pd加上新的平均值
        newEMAvg = pd.concat([newEM,pd.DataFrame(newEM.mean()).rename(columns = {0:"平均"}).T])

        EMModel = header(newEMAvg)
        w.EMTableView_2.setModel(EMModel)

        # 畫直方圖 -> 彈性模數
        if True not in np.isnan(list(newEMAvg.iloc[-1])):
            EMAvgList = list(newEMAvg.iloc[-1])
            plt.figure(figsize=(5,3))
            B = plt.bar(exRatio, EMAvgList, align='center', alpha=0.5)
            # plt.xticks(y_pos, objects)
            plt.ylabel('Elastic modulus(kgf/cm2)');
            EMTitleName = w.EMTitle.text()
            if EMTitleName == "":
                EMTitleName = 'Elastic modulus'
            plt.title(EMTitleName)

            createLabels(B)
            plt.savefig(path+'\\Elastic modulus-剔除歧異值.png',bbox_inches ='tight');

            figName = path+"\\Elastic modulus-剔除歧異值.png"
            layout_Figure(lab5,figName,scene5,w.graphicsView_5,scale=[1,1])
            
            # ToExcel
            ExcelFileName = path+'\\彈性模數-歧異值.xlsx'
#             workbook = xlsxwriter.Workbook(ExcelFileName)     #新建文件
#             worksheet1 = workbook.add_worksheet('Elastic Modulus')             #新建sheet
#             bold = workbook.add_format({'bold': True})
            newEMAvg = newEMAvg.fillna("剔除掉")
            newEMAvg.to_excel(ExcelFileName)
#             worksheet1.insert_image(1,10, path+'\\Elastic modulus-剔除歧異值.png',{'x_scale':1,'y_scale':1})
#             workbook.close()  # 保存並關閉

#             from openpyxl import load_workbook

#             book = load_workbook(ExcelFileName)
#             writer = pd.ExcelWriter(ExcelFileName, engine='openpyxl') 
#             writer.book = book
#             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

#             newEMAvg.to_excel(writer, "Elastic Modulus")
            
            QMessageBox.information(w,'Congratulation', "剔除後的彈模數據成功存成Excel~!")
            plt.close("all")
            
        else:
            QMessageBox.critical(w,'Error', "該剔除值下，有配比已經無資料!")
    except:
        QMessageBox.critical(w,'Error', "尚未有彈模數據!")



def plot():
    
    # python中，def裡面的全域變數需要先宣告
    global exRatio,exNum,area,PathFile,Savefile,fileName
    
    #判斷是否有選擇text檔案
    if w.PathTxt.toPlainText() == "":
        QMessageBox.critical(w,'Error', "沒有選檔案!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return
    
    #判斷是否有勾選cube or cylinder
    if w.Cylinder.isChecked():
        area = np.pi/4*10**2
        
    elif w.Cube.isChecked():
        area = 5**2
        
    else:
        QMessageBox.critical(w,'Error', "沒有選試體尺寸!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return
    
    #判斷是否有輸入配比數量
    if w.inputWidget.rowCount() == 0:
        QMessageBox.critical(w,'Error', "沒有輸入配比數量!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return
    
    
    #判斷是否有選擇存檔路徑
    if w.savePathTxt.toPlainText() == "":
        QMessageBox.critical(w,'Error', "沒有選存檔路徑!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return
    
    #判斷是否有輸入存檔檔名
    if w.saveFileName.text() == "":
        QMessageBox.critical(w,'Error', "沒有輸入存檔名稱!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return
    

    scene = QGraphicsScene()
    
    Savefile = w.saveFileName.text()
    
    exRatio = []
    exNum = []
    
    
        
    for i in range(w.inputWidget.rowCount()):
        
        # 配比種類
        if w.inputWidget.item(i,0) == None:
            QMessageBox.critical(w,'Error', "配比種類沒有完整輸入!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
            return
        exName = w.inputWidget.item(i,0).text()
        exRatio.append(exName)

        # 配比數量
        try:
            if float(w.inputWidget.item(i,1).text()) == 0:
                QMessageBox.critical(w,'Error', "配比數量不能為零!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
                return
            if float(w.inputWidget.item(i,1).text()) != int(float(w.inputWidget.item(i,1).text())):
                QMessageBox.critical(w,'Error', "配比數量不能為小數!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
                return
            if float(w.inputWidget.item(i,1).text()) <0 :
                QMessageBox.critical(w,'Error', "配比數量必須為正整數!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
                return
        except:
            QMessageBox.critical(w,'Error', "配比數量必須為數字，不能為文字!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
            return
        
        exNum.append(float(w.inputWidget.item(i,1).text()))
        
    if len(set(exRatio)) != len(exRatio):
        QMessageBox.critical(w,'Error', "配比種類名稱重複!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return

    with open('{}'.format(fileName),'r',encoding="utf-8") as f:  # with open 語法區域內執行完畢會自動釋放資源
        fileRead = f.read().split('\n\n')
    fileRead.pop(0)
    
    global n
    n = len(fileRead)
    if sum(exNum) != n:
        QMessageBox.critical(w,'Error', "試體個數不符合，請重新確認!", QMessageBox.Yes | QMessageBox.No,QMessageBox.Yes)
        return

    displDict = {}
    forceDict = {}
    stressDict = {}
    strainDict = {}
    Stress2StrainDict = {}
    
    # 處理txt檔
    for i in range(len(fileRead)):
        eachCountDatas = fileRead[i].split('segments\n')[-1]
        eachLineDataList = eachCountDatas.split('\n')

        displList = []
        forceList = []
        stressList = []
        strainList = []

        for j in range(len(eachLineDataList)):
            oneLineDataList = eachLineDataList[j].split('\t')
            if oneLineDataList == ['']:
                break

            displ = float(oneLineDataList[1])
            force = float(oneLineDataList[2])
            D1 = float(oneLineDataList[3])
            D2 = float(oneLineDataList[4])
            strain = abs((D1+D2)/2)/100*5.384        # 除以夾具距離(100mm)  # LVDT1mm = 實際5.384mm
    #         strain = (D1+D2)/200*10 
            stress = force/9.81/area                 # 除以截面積(kgf/cm2)

            forceList.append(force)
            displList.append(displ)
            stressList.append(stress)  
            strainList.append(strain)           

        displDict['data'+str(i+1)] = displList
        forceDict['data'+str(i+1)] = forceList
        stressDict['data'+str(i+1)] = stressList
        strainDict['data'+str(i+1)] = strainList
    

    # Target: # ex = ['0%','0%','0%','0.25%','0.25%','0.25%','0.5%','0.5%','0.5%','1.0%','1.0%','1.0%','1.5%','1.5%','1.5%','2%','2%','2%']
    ex1 = []  
    ex = []
    for j in range(len(exRatio)):
        ex1.append(exRatio[j:j+1])
    for j in range(len(ex1)):
        ex+=ex1[j]*int(exNum[j])
        
    # 配比個數分段點
    exNum = [0]+exNum
    for i in  range(1,len(exNum)):
        exNum[i] = int(exNum[i])+int(exNum[i-1])
    
    def Process_NoEM(n,stressDict,strainDict,PathFile,Savefile,exRatio):
        maxStressDict = {}
        maxStressList = []
        step = 0
        
        for i in range(n):
            Stress_i_List = stressDict['data'+str(i+1)]
            Strain_i_List = strainDict['data'+str(i+1)]
            # 裝入max
            maxStressList.append(max(Stress_i_List))
            for j in range(len(exNum)):
                if i+1 == int(exNum[j]):
                    maxStressDict[ex[i]] = maxStressList[exNum[j-1]:exNum[j]]
            step +=100/n
            w.progressBar.setValue(round(step))
                    
        # DataFrame
        comp = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in maxStressDict.items()]))
        comp = comp.rename(index={0: "第一顆", 1: "第二顆", 2: "第三顆", 3: "第四顆", 4: "第五顆(以下以此類推)"})
        comp = pd.concat([comp,pd.DataFrame(comp.mean()).rename(columns = {0:"平均"}).T])
        
        compAvgList = []
        for i in range(len(exRatio)):
            compAvgList.append(round(np.average(maxStressDict[exRatio[i]]),3))
       
        # 建立資料夾/指定存檔路徑
        global path
        path = os.path.abspath(PathFile)+"\\"+Savefile
        if not os.path.isdir(path):
            os.mkdir(path)
    
        # 畫直方圖 -> 抗壓強度
        plt.figure(figsize=(5,3))
        A = plt.bar(exRatio, compAvgList, align='center', alpha=0.5)
        # plt.xticks(y_pos, objects)
        plt.ylabel('Compressive Strength(kgf/cm2)');
        compTitleName = w.compTitle.text()
        if compTitleName == "":
            compTitleName = 'Compressive Strength'
        plt.title(compTitleName)

        createLabels(A)
        plt.savefig(path+'\\Compressive Strength.png',bbox_inches ='tight');
        
        # graphicsView_2 -> 放抗壓預覽圖
        scene2 = QGraphicsScene()
        lab2 = QLabel()
        figName = path+"\\Compressive Strength.png"
        layout_Figure(lab2,figName,scene2,w.graphicsView_2,scale = [1,1])
        
        # graphicsView_4 -> 放比較大的抗壓圖
        global scene4,lab4
        scene4 = QGraphicsScene()
        lab4 = QLabel()
        figName = path+"\\Compressive Strength.png"
        layout_Figure(lab4,figName,scene4,w.graphicsView_4,scale = [1,1])
        
        # ToExcel
        ExcelFileName = path+'\\AllData.xlsx'
        workbook = xlsxwriter.Workbook(ExcelFileName)     #新建文件
        worksheet1 = workbook.add_worksheet('Compressive Strength')             #新建sheet
        bold = workbook.add_format({'bold': True})
        worksheet1.insert_image(1,10, path+'\\Compressive Strength.png',{'x_scale':1,'y_scale':1})
        workbook.close()  # 保存並關閉

        from openpyxl import load_workbook

        book = load_workbook(ExcelFileName)
        writer = pd.ExcelWriter(ExcelFileName, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        comp.to_excel(writer, "Compressive Strength")

        compModel = header(comp)
        w.compTableView.setModel(compModel)
        w.EMTableView.setModel(None)
        
        w.ReEMPlotBtn.setEnabled(False)
        
        writer.save()
        QMessageBox.information(w,'Congratulation', "成功存成Excel~!")
        
    def Process_withEM(n,stressDict,strainDict,PathFile,Savefile,exRatio):
        
        errorList = []
        maxStressDict = {}
        maxStressList = []
        EMList = []
        abList = []
        EMDict = {}
        step = 0
        
        for i in range(n):

            Stress_i_List = stressDict['data'+str(i+1)]
            Strain_i_List = strainDict['data'+str(i+1)]

            # 找應變0.00005到MAX應力的0.45之間(兩個點)
            UpperStress = max(Stress_i_List)*0.45

            for j in range(len(Stress_i_List)):
                if Stress_i_List[j] >= UpperStress:
                    UpperPoint = Strain_i_List[j-1]
                    UpperPointIndex = j-1
                    break

            for j in range(len(Strain_i_List)):
                if Strain_i_List[j] >= 0.0002:
                    LowerPoint = Strain_i_List[j]
                    LowerPointIndex = j
                    break

            if LowerPointIndex > UpperPointIndex:
                a = LowerPointIndex
                LowerPointIndex = UpperPointIndex
                UpperPointIndex = a

            # 迴歸資料
            x = np.array(Strain_i_List).reshape(-1,1)
            y = np.array(Stress_i_List).reshape(-1,1)
            xmax = np.array(Strain_i_List[LowerPointIndex:UpperPointIndex]).reshape(-1,1)
            ymax = np.array(Stress_i_List[LowerPointIndex:UpperPointIndex]).reshape(-1,1)

            xmaxcf = Strain_i_List[LowerPointIndex:UpperPointIndex]
            ymaxcf = Stress_i_List[LowerPointIndex:UpperPointIndex]
            
            # 回歸錯誤的除錯
            try:
                popt, pcov = curve_fit(func, xmaxcf, ymaxcf)
            except:

                for j in range(len(exRatio)):
                    if exRatio.index(ex[i]) == j:
                        errorList.append([ex[i],str(i-exNum[j]+1)])
                pass

            abList.append([popt[0],popt[1]])
            
            # 裝入max
            maxStressList.append(max(Stress_i_List))
            EMList.append(popt[0])

            for j in range(len(exNum)):
                if i+1 == int(exNum[j]):
                    maxStressDict[ex[i]] = maxStressList[exNum[j-1]:exNum[j]]
                    EMDict[ex[i]] = EMList[exNum[j-1]:exNum[j]]
            
            # 建立資料夾/指定存檔路徑
            global path
            path = os.path.abspath(PathFile)+"\\"+Savefile
            if not os.path.isdir(path):
                os.mkdir(path)

            # 繪小圖 -> 存檔用
            fig1, (ax1, ax2) = plt.subplots(1,2,figsize = (15,5))
            axesPlot(ax1,ax2,x,y,i,Strain_i_List,ex,xmax,ymax,popt)
            fig1.savefig(path+'\\fig_'+str(i+1)+'.png')

            # 繪大圖 -> 預覽用
            axesPlot(ax[i,0],ax[i,1],x,y,i,Strain_i_List,ex,xmax,ymax,popt)
            
            step +=100/n
            w.progressBar.setValue(round(step))
        fig2.savefig(path+'\\fig_All.png',dpi=100)
        
        # 回歸error
        if len(errorList)!=0:
            for each in errorList:
                QMessageBox.critical(w,'Error', "配比 "+each[0]+" 的第"+each[1]+"個數據無法回歸!")

        # 將大圖放到第一張預覽圖上面
        lab1 = QLabel()
        figName = path+"\\fig_All.png"
        layout_Figure(lab1,figName,scene,w.graphicsView,scale = [1,1])
        
        # DataFrame
        global comp,comp_avg,EM,EM_avg
        
        comp = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in maxStressDict.items()]))
        comp = comp.rename(index={0: "第一顆", 1: "第二顆", 2: "第三顆", 3: "第四顆", 4: "第五顆(以下以此類推)"})
        comp_avg = pd.concat([comp,pd.DataFrame(comp.mean()).rename(columns = {0:"平均"}).T])

        EM = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in EMDict.items()]))
        EM = EM.rename(index={0: "第一顆", 1: "第二顆", 2: "第三顆", 3: "第四顆", 4: "第五顆(以下以此類推)"})
        EM_avg = pd.concat([EM[EM>0],pd.DataFrame(EM[EM>0].mean()).rename(columns = {0:"平均"}).T])

        compAvgList = []
        for i in range(len(exRatio)):
            compAvgList.append(round(np.average(maxStressDict[exRatio[i]]),3))

        EMAvgList = []
        for i in range(len(exRatio)):
            EMAvgList.append(round(np.average(EMDict[exRatio[i]]),3))


        plt.clf() # 清空圖

        # 畫直方圖 -> 抗壓強度
        plt.figure(figsize=(5,3))
        A = plt.bar(exRatio, compAvgList, align='center', alpha=0.5)
        # plt.xticks(y_pos, objects)
        plt.ylabel('Compressive Strength(kgf/cm2)');
        compTitleName = w.compTitle.text()
        if compTitleName == "":
            compTitleName = 'Compressive Strength'
        plt.title(compTitleName)

        createLabels(A)
        plt.savefig(path+'\\Compressive Strength.png',bbox_inches ='tight');

        # 畫直方圖 -> 彈性模數
        plt.figure(figsize=(5,3))
        B = plt.bar(exRatio, EMAvgList, align='center', alpha=0.5)
        # plt.xticks(y_pos, objects)
        plt.ylabel('Elastic modulus(kgf/cm2)');
        EMTitleName = w.EMTitle.text()
        if EMTitleName == "":
            EMTitleName = 'Elastic modulus'
        plt.title(EMTitleName)

        createLabels(B)
        plt.savefig(path+'\\Elastic modulus.png',bbox_inches ='tight');


        # 將抗壓放到第二張預覽圖上面
        scene2 = QGraphicsScene();
        lab2 = QLabel()
        figName = path+"\\Compressive Strength.png"
        layout_Figure(lab2,figName,scene2,w.graphicsView_2,scale = [1,1])
        
        # graphicsView_4 -> 放比較大的圖
        global scene4,lab4
        scene4 = QGraphicsScene()
        lab4 = QLabel()
        figName = path+"\\Compressive Strength.png"
        layout_Figure(lab4,figName,scene4,w.graphicsView_4,scale=[1,1])
        
        # 將彈模放到第三張預覽圖上面
        scene3 = QGraphicsScene()
        lab3 = QLabel()
        figName = path+"\\Elastic modulus.png"
        layout_Figure(lab3,figName,scene3,w.graphicsView_3,scale = [1,1])
        
        # graphicsView_5 -> 放比較大的圖
        global scene5,lab5
        scene5 = QGraphicsScene()
        lab5 = QLabel()
        figName = path+"\\Elastic modulus.png"
        layout_Figure(lab5,figName,scene5,w.graphicsView_5,scale = [1,1])
        
        # ToExcel
        ExcelFileName = path+'\\AllData.xlsx'
        workbook = xlsxwriter.Workbook(ExcelFileName)     #新建文件
        worksheet1 = workbook.add_worksheet('All Data')             #新建sheet
        bold = workbook.add_format({'bold': True})

        for i in range(len(fileRead)):

            worksheet1.write(20*i+1,0,'第'+str(i+1)+'筆Data: '+ex[i])  

            worksheet1.write(20*i+2,0,'最大抗壓強度f"c(kgf/cm2):\t')

            worksheet1.write(20*i+2,1,max(stressDict['data'+str(i+1)]))

            worksheet1.write(20*i+3,0,'理論楊氏係數(kgf/cm2):\t')

            worksheet1.write(20*i+3,1,12000*(max(stressDict['data'+str(i+1)]))**0.5)

            worksheet1.write(20*i+4,0,'回歸係數(y=ax+b),[a,b]=\t')

            worksheet1.write(20*i+4,1,str(abList[i]))

            worksheet1.write(20*i+5,0,'\n')

            worksheet1.insert_image(20*i,6, path+'\\fig_'+str(i+1)+'.png',{'x_scale':0.65,'y_scale':0.65})  


        worksheet2 = workbook.add_worksheet('Compressive Strength')             #新建sheet
        worksheet2.insert_image(1,10, path+'\\Compressive Strength.png',{'x_scale':1,'y_scale':1})

        worksheet3 = workbook.add_worksheet('Elastic Modulus')             #新建sheet
        worksheet3.insert_image(1,10, path+'\\Elastic modulus.png',{'x_scale':1,'y_scale':1})

        workbook.close()  # 保存並關閉

        from openpyxl import load_workbook

        book = load_workbook(ExcelFileName)
        writer = pd.ExcelWriter(ExcelFileName, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        comp_avg.to_excel(writer, "Compressive Strength")
        EM_avg.to_excel(writer, "Elastic Modulus")
        
        # 修改TableView的header 
        compModel = header(comp_avg)
        EMModel = header(EM_avg)
        w.compTableView.setModel(compModel)
        w.EMTableView.setModel(EMModel)
        
        w.ReEMPlotBtn.setEnabled(True)

        writer.save()
        QMessageBox.information(w,'Congratulation', "成功存成Excel~!")
        
    
    # 如果勾選Cube -> 不要畫圖
    if w.Cube.isChecked():
        
        lab1 = QLabel()
        lab1.setText("No Figure (No LVDT)")
        lab1.setStyleSheet("QLabel{color:rgb(0,0,0,255);background-color:rgb(255,255,255,255);font-size:50px;font-weight:normal;font-family:Arial;}")
        lab1.setScaledContents(True)
        scene.addWidget(lab1)
        w.graphicsView.setScene(scene)
        w.graphicsView_3.setScene(scene)
        w.graphicsView_5.setScene(scene)
        
        Process_NoEM(n,stressDict,strainDict,PathFile,Savefile,exRatio)
        
    else:
        # 建立大圖以做subplots
        fig2, ax = plt.subplots(n,2,figsize = (11.5,3.3*n))
        plt.tight_layout(pad=3.5)
        
        Process_withEM(n,stressDict,strainDict,PathFile,Savefile,exRatio)
        
    # 清空篩選的table與歸零垂直棒
    plt.close("all")
    w.compTableView_2.setModel(None)
    w.EMTableView_2.setModel(None)
    w.compVerticalSlider.setValue(20)
    w.EMVerticalSlider.setValue(20)
    
# 開啟 UI_Designer
app = QApplication(sys.argv)
w = loadUi('comp3.ui')

# 預設 graphicView1,2,3上面的文字
scene1 = QGraphicsScene();scene2 = QGraphicsScene();scene3 = QGraphicsScene()
label1 = QLabel();label2 = QLabel();label3 = QLabel()

text = "Stress-Strain Figure"
Initial_layout_Figure(label1,text,scene1,w.graphicsView)
text = "Compressive Strength Figure"
Initial_layout_Figure(label2,text,scene2,w.graphicsView_2)
text = "Elastic Modulus Figure"
Initial_layout_Figure(label3,text,scene3,w.graphicsView_3)

w.setWindowTitle('MTSDataProcess')

# 進度條
w.progressBar.setMinimum(0)
w.progressBar.setMaximum(100)
w.progressBar.setValue(0)

# button功能
w.selectBtn.clicked.connect(read)
w.spinBox.valueChanged.connect(change_row)
w.saveBtn.clicked.connect(save)
w.pltBtn.clicked.connect(plot)

# 垂直滑動棒 -> 控制剔除的%數，預設20%
w.compVerticalSlider.setValue(20)
w.EMVerticalSlider.setValue(20)

w.compVerticalSlider.valueChanged.connect(compshow)
w.EMVerticalSlider.valueChanged.connect(EMshow)

w.ReCompPlotBtn.clicked.connect(compRePlot)
w.ReEMPlotBtn.clicked.connect(EMRePlot)


w.show()
app.exec_()






